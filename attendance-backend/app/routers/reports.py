from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
import datetime
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.database import get_db
from app.models.student import Student
from app.models.user import User
from app.models.attendance import AttendanceRecord, AttendanceSummary
from app.models.course import CourseEnrollment, Course
from app.utils.security import get_current_user

router = APIRouter(prefix="/api/reports", tags=["Reports"])

@router.get("/download/pdf/{student_id}")
async def download_attendance_pdf(student_id: str, db: Session = Depends(get_db)):
    # 1. Fetch Student and User Details
    student = db.query(Student).filter(Student.student_id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    user = db.query(User).filter(User.user_id == student.user_id).first()
    
    # 2. Fetch Attendance Records across all courses
    enrollments = db.query(CourseEnrollment).filter(CourseEnrollment.student_id == student_id).all()
    enrollment_ids = [e.enrollment_id for e in enrollments]
    
    records = db.query(AttendanceRecord, Course.course_name, Course.course_code)\
        .join(CourseEnrollment, AttendanceRecord.enrollment_id == CourseEnrollment.enrollment_id)\
        .join(Course, CourseEnrollment.course_id == Course.course_id)\
        .filter(AttendanceRecord.enrollment_id.in_(enrollment_ids))\
        .order_by(AttendanceRecord.class_date.desc()).all()

    # 3. Create PDF
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Header
    p.setFont("Helvetica-Bold", 18)
    p.drawString(100, height - 50, "AttendLink - Attendance Report")
    
    p.setFont("Helvetica", 12)
    p.drawString(100, height - 80, f"Name: {user.full_name}")
    p.drawString(100, height - 95, f"Roll Number: {student.roll_number}")
    p.drawString(100, height - 110, f"Generated On: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    # Table Header
    p.line(100, height - 130, 500, height - 130)
    p.setFont("Helvetica-Bold", 11)
    p.drawString(100, height - 145, "Date")
    p.drawString(200, height - 145, "Course")
    p.drawString(400, height - 145, "Status")
    p.line(100, height - 150, 500, height - 150)
    
    # Rows
    y = height - 165
    p.setFont("Helvetica", 10)
    for record, course_name, course_code in records:
        if y < 50:
            p.showPage()
            y = height - 50
            p.setFont("Helvetica", 10)
            
        p.drawString(100, y, str(record.class_date))
        p.drawString(200, y, f"{course_code} - {course_name[:25]}")
        
        status_color = colors.green if record.status == 'present' else colors.red
        p.setFillColor(status_color)
        p.drawString(400, y, record.status.upper())
        p.setFillColor(colors.black)
        
        y -= 20

    p.showPage()
    p.save()
    
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=attendance_{student.roll_number}.pdf"})

@router.get("/download/excel/{student_id}")
async def download_attendance_excel(student_id: str, db: Session = Depends(get_db)):
    # 1. Fetch Data
    student = db.query(Student).filter(Student.student_id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    user = db.query(User).filter(User.user_id == student.user_id).first()
    enrollments = db.query(CourseEnrollment).filter(CourseEnrollment.student_id == student_id).all()
    enrollment_ids = [e.enrollment_id for e in enrollments]
    
    records = db.query(AttendanceRecord, Course.course_name, Course.course_code)\
        .join(CourseEnrollment, AttendanceRecord.enrollment_id == CourseEnrollment.enrollment_id)\
        .join(Course, CourseEnrollment.course_id == Course.course_id)\
        .filter(AttendanceRecord.enrollment_id.in_(enrollment_ids))\
        .order_by(AttendanceRecord.class_date.desc()).all()

    # 2. Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Header Styling
    header_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal="center")

    # Metadata
    ws["A1"] = "Attendance Report"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Student: {user.full_name} ({student.roll_number})"
    ws["A3"] = f"Exported: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"

    # Table Header
    headers = ["Date", "Course Code", "Course Name", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align

    # Data Rows
    for row_idx, (record, course_name, course_code) in enumerate(records, 6):
        ws.cell(row=row_idx, column=1, value=str(record.class_date))
        ws.cell(row=row_idx, column=2, value=course_code)
        ws.cell(row=row_idx, column=3, value=course_name)
        ws.cell(row=row_idx, column=4, value=record.status.upper())

    # Column Widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=attendance_{student.roll_number}.xlsx"}
    )
